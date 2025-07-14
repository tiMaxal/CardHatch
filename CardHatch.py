"""
Install Required Packages:
 bash
pip install pandas reportlab openpyxl

"""

"""
Flashcard PDF Generator

This script creates a Tkinter-based GUI application that
 generates a double-sided flashcard PDF from a CSV, Excel, or ODS file.
The application allows users to select an input file 
 (showing CSV, Excel, and ODS files simultaneously by default), specify column
  names for front and back content (with an option to autofill from the first two columns),
    card layout, page size, margins, color bars, flip mode for duplex printing, font size,
     font family (including Arial), font style, text color, background color, variable quantities via a 'qty' column,
      and a global quantity multiplier limited to 4 digits. Cards and text are centered,
        and front/back pages are aligned for duplex printing.
The GUI is organized into sections, is resizable, and includes a vertical scrollbar on the right
 to access all elements if they exceed the window height. The 'Colors' and 'Font Details' sections
  are split into two equal parts for front and back, with options for text color, top/bottom color bars,
   background color, font size, family, and style for each side.
All elements, including Start and Exit buttons, are visible on window open.
Settings are saved to a JSON file for persistence.
Comprehensive logging tracks all stages.
Empty, non-numeric, non-integer, or non-positive 'qty' values default to 1, with user notification.
Text in cells can include carriage returns (CR, LF, or CRLF) for multi-line formatting,
 respected in the output PDF.

Dependencies:
- pandas
- reportlab
- openpyxl
- odfpy
- csv (standard library)

Install requirements: pip install pandas reportlab openpyxl odfpy
"""

import tkinter as tk
from tkinter import filedialog, messagebox, colorchooser
import json
import os
import math
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor
from reportlab.pdfbase.pdfmetrics import stringWidth
import logging
import csv
import openpyxl
import odf.opendocument
from odf.table import Table
from odf.text import P
import io

# ---------- Logging Setup ----------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("cardhatch.log"),
    ],
)
logger = logging.getLogger(__name__)

# ---------- Settings Management ----------

DEFAULT_SETTINGS = {
    "file_path": "",
    "output_file": "",
    "front_column": "",
    "back_column": "",
    "cards_per_row": 3,
    "card_width": 85.6,
    "card_height": 55.0,
    "page_size": "210x297",
    "margins": "10,10,10,10",
    "front_color_bar_top": False,
    "front_color_bar_bottom": False,
    "front_color_bar_top_color": "#FF0000",
    "front_color_bar_bottom_color": "#0000FF",
    "front_background_color": "#FFFFFF",
    "back_color_bar_top": False,
    "back_color_bar_bottom": False,
    "back_color_bar_top_color": "#FF0000",
    "back_color_bar_bottom_color": "#0000FF",
    "back_background_color": "#FFFFFF",
    "truncate": False,
    "flip_mode": "long",
    "front_font_size": 12,
    "front_font_family": "Helvetica",
    "front_font_style": "Normal",
    "front_text_color": "#000000",
    "back_font_size": 12,
    "back_font_family": "Helvetica",
    "back_font_style": "Normal",
    "back_text_color": "#000000",
    "use_qty_column": True,
    "quantity_multiplier": 1,
    "autofill_columns": False,
}

SETTINGS_FILE = "cardhatch_settings.json"


def load_settings():
    """
    Load settings from a JSON file or return default settings if the file doesn't exist.

    Returns:
        dict: Loaded or default settings.
    """
    logger.info(f"Loading settings from {SETTINGS_FILE}")
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, "r") as f:
                settings = json.load(f)
                logger.info("Settings loaded successfully")
                return settings
        else:
            logger.info("Settings file not found, using default settings")
            return DEFAULT_SETTINGS.copy()
    except Exception as e:
        logger.error(f"Failed to load settings: {e}")
        return DEFAULT_SETTINGS.copy()


settings = load_settings()

# ---------- Utility Functions ----------


def mm(val):
    """
    Convert millimeters to PDF points (1 mm = 2.83465 points).

    Args:
        val (float): Value in millimeters.

    Returns:
        float: Value in PDF points.
    """
    return float(val) * 2.83465


def wrap_text(text, font_name, font_size, max_width_pt, max_lines, truncate=False):
    """
    Wrap text to fit within a maximum width and number of lines, respecting explicit newline characters.

    Args:
        text (str): Text to wrap, may contain \n or \r\n.
        font_name (str): Font name for text measurement.
        font_size (float): Font size in points.
        max_width_pt (float): Maximum width in points.
        max_lines (int): Maximum number of lines.
        truncate (bool): If True, truncate text with ellipsis.

    Returns:
        tuple: (list of wrapped lines, bool indicating if text overflowed).
    """
    logger.debug(
        f"Wrapping text: {text[:50]}... (max_width={max_width_pt}, max_lines={max_lines})"
    )
    text = str(text).replace("\r\n", "\n")
    if "\n" not in text and len(text) > 50:
        logger.warning(
            "No newlines detected in long text; ensure fields are quoted with literal newlines"
        )
    explicit_lines = text.split("\n")
    final_lines = []
    overflowed = False

    for line in explicit_lines:
        if not line.strip():
            final_lines.append("")
            continue
        words = line.split()
        current_line = ""
        for word in words:
            test_line = (current_line + " " + word).strip()
            if stringWidth(test_line, font_name, font_size) <= max_width_pt:
                current_line = test_line
            else:
                if current_line:
                    final_lines.append(current_line)
                    if len(final_lines) >= max_lines:
                        if truncate:
                            final_lines[-1] = (
                                final_lines[-1][: int(max_width_pt / font_size)] + "..."
                            )
                            return final_lines[:max_lines], False
                        return final_lines[:max_lines], True
                current_line = word
        if current_line:
            final_lines.append(current_line)
            if len(final_lines) >= max_lines:
                if truncate:
                    final_lines[-1] = (
                        final_lines[-1][: int(max_width_pt / font_size)] + "..."
                    )
                    return final_lines[:max_lines], False
                return final_lines[:max_lines], True

    if len(final_lines) > max_lines:
        if truncate:
            final_lines = final_lines[:max_lines]
            final_lines[-1] = final_lines[-1][: int(max_width_pt / font_size)] + "..."
            return final_lines, False
        return final_lines[:max_lines], True

    return final_lines, False


def draw_cut_lines(
    c,
    page_w_pt,
    page_h_pt,
    offset_left_pt,
    offset_top_pt,
    card_width_pt,
    card_height_pt,
    cards_per_row,
    cards_per_col,
):
    """
    Draw cut lines on the PDF for card separation.

    Args:
        c (Canvas): ReportLab canvas object.
        page_w_pt (float): Page width in points.
        page_h_pt (float): Page height in points.
        offset_left_pt (float): Left offset for card grid in points.
        offset_top_pt (float): Top offset for card grid in points.
        card_width_pt (float): Card width in points.
        card_height_pt (float): Card height in points.
        cards_per_row (int): Number of cards per row.
        cards_per_col (int): Number of cards per column.
    """
    logger.debug(f"Drawing cut lines for {cards_per_row}x{cards_per_col} grid")
    c.setStrokeColor(HexColor("#888888"))
    c.setLineWidth(0.5)
    for i in range(cards_per_row + 1):
        x = offset_left_pt + i * card_width_pt
        c.line(x, offset_top_pt, x, page_h_pt - offset_top_pt)
    for j in range(cards_per_col + 1):
        y = page_h_pt - offset_top_pt - j * card_height_pt
        c.line(offset_left_pt, y, page_w_pt - offset_left_pt, y)


def reorder_for_back(page_indices, cards_per_row, cards_per_col, flip_mode):
    """
    Reorder card indices for the back page based on flip mode for duplex printing.

    Args:
        page_indices (list): List of card indices for the page.
        cards_per_row (int): Number of cards per row.
        cards_per_col (int): Number of cards per column.
        flip_mode (str): 'long' or 'short' edge flip.

    Returns:
        list: Reordered indices for the back page.
    """
    logger.debug(f"Reordering indices for back page with flip_mode={flip_mode}")
    reordered = []
    if flip_mode == "long":
        for r in range(cards_per_col):
            row = page_indices[r * cards_per_row : (r + 1) * cards_per_row]
            reordered.extend(row[::-1])
    else:
        for r in reversed(range(cards_per_col)):
            row = page_indices[r * cards_per_row : (r + 1) * cards_per_row]
            reordered.extend(row)
    return reordered


# ---------- Main Application ----------


class FlashcardApp(tk.Tk):
    """
    Tkinter GUI application for generating printable, double-sided flashcard PDFs from spreadsheet data.

    Provides fields for input file selection, column names (with autofill option), card layout, page size, margins,
    color bars, flip mode, font size, font family, font style, text color, background color, variable quantities, and
    a global quantity multiplier (limited to 4 digits). The 'Colors' and 'Font Details' sections are split into front
    and back subsections. GUI is organized into sections, is resizable, and includes a vertical scrollbar.
    All elements, including Start and Exit buttons, are visible on window open. Cards and text are centered, and
    front/back pages are aligned for duplex printing.
    """

    def __init__(self):
        """Initialize the GUI with input fields prefilled from saved or default settings, organized into sections."""
        super().__init__()
        logger.info("Initializing FlashcardApp GUI")
        self.title("Flashcard PDF Generator")
        self.geometry("750x850")  # Increased height again to ensure buttons visibility
        self.resizable(True, True)  # Enable window resizing

        # Create canvas and scrollbar
        self.canvas = tk.Canvas(self, highlightthickness=0)
        self.scrollbar = tk.Scrollbar(
            self, orient=tk.VERTICAL, command=self.canvas.yview
        )
        self.scrollable_frame = tk.Frame(self.canvas)

        # Configure canvas scrolling
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Create a window in the canvas for the scrollable frame
        self.canvas_frame = self.canvas.create_window(
            (0, 0), window=self.scrollable_frame, anchor="nw"
        )

        # Bind canvas resizing and scrolling
        self.scrollable_frame.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        # Enable mouse wheel scrolling
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)  # Windows
        self.canvas.bind_all("<Button-4>", self._on_mousewheel)  # Linux
        self.canvas.bind_all("<Button-5>", self._on_mousewheel)  # Linux

        # Main container inside scrollable frame
        frame_main = tk.Frame(self.scrollable_frame)
        frame_main.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        frame_main.columnconfigure(0, weight=1)
        frame_main.rowconfigure(0, weight=1)
        frame_main.rowconfigure(1, weight=1)
        frame_main.rowconfigure(2, weight=1)
        frame_main.rowconfigure(3, weight=1)
        frame_main.rowconfigure(4, weight=1)

        # Section: File Paths & Column Names
        frame_file = tk.LabelFrame(
            frame_main, text="File Paths & Column Names", padx=5, pady=5
        )
        frame_file.grid(row=0, column=0, sticky=tk.EW, pady=5)
        frame_file.columnconfigure(1, weight=1)

        tk.Label(frame_file, text="CSV/Excel/ODS File Path:").grid(
            row=0, column=0, sticky=tk.W
        )
        self.entry_file = tk.Entry(frame_file)
        self.entry_file.grid(row=0, column=1, sticky=tk.EW)
        self.entry_file.insert(0, settings.get("file_path", ""))
        btn_browse = tk.Button(frame_file, text="Browse...", command=self.browse_file)
        btn_browse.grid(row=0, column=2)

        tk.Label(frame_file, text="Output PDF File:").grid(row=1, column=0, sticky=tk.W)
        self.entry_output = tk.Entry(frame_file)
        self.entry_output.grid(row=1, column=1, sticky=tk.EW)
        default_output = (
            os.path.splitext(settings.get("file_path", ""))[0] + ".pdf"
            if settings.get("file_path", "")
            else "flashcards.pdf"
        )
        self.entry_output.insert(0, settings.get("output_file", default_output))
        btn_output_browse = tk.Button(
            frame_file, text="Browse...", command=self.browse_output_file
        )
        btn_output_browse.grid(row=1, column=2)

        self.autofill_columns_var = tk.BooleanVar(
            value=settings.get("autofill_columns", False)
        )
        tk.Checkbutton(
            frame_file,
            text="Autofill Column Names from File",
            variable=self.autofill_columns_var,
            command=self.toggle_column_entries,
        ).grid(row=2, column=0, columnspan=3, sticky=tk.W)

        tk.Label(frame_file, text="Front Column Name:").grid(
            row=3, column=0, sticky=tk.W
        )
        self.entry_front = tk.Entry(frame_file)
        self.entry_front.grid(row=3, column=1, sticky=tk.EW)
        self.entry_front.insert(0, settings.get("front_column", ""))

        tk.Label(frame_file, text="Back Column Name:").grid(
            row=4, column=0, sticky=tk.W
        )
        self.entry_back = tk.Entry(frame_file)
        self.entry_back.grid(row=4, column=1, sticky=tk.EW)
        self.entry_back.insert(0, settings.get("back_column", ""))

        # Section: Card Layout
        frame_layout = tk.LabelFrame(frame_main, text="Card Layout", padx=5, pady=5)
        frame_layout.grid(row=1, column=0, sticky=tk.EW, pady=5)
        frame_layout.columnconfigure(1, weight=1)

        tk.Label(frame_layout, text="Cards per Row:").grid(row=0, column=0, sticky=tk.W)
        self.entry_cards_per_row = tk.Entry(frame_layout)
        self.entry_cards_per_row.grid(row=0, column=1, sticky=tk.EW)
        self.entry_cards_per_row.insert(0, str(settings.get("cards_per_row", 3)))

        tk.Label(frame_layout, text="Card Width (mm):").grid(
            row=1, column=0, sticky=tk.W
        )
        self.entry_card_width = tk.Entry(frame_layout)
        self.entry_card_width.grid(row=1, column=1, sticky=tk.EW)
        self.entry_card_width.insert(0, str(settings.get("card_width", 85.6)))

        tk.Label(frame_layout, text="Card Height (mm):").grid(
            row=2, column=0, sticky=tk.W
        )
        self.entry_card_height = tk.Entry(frame_layout)
        self.entry_card_height.grid(row=2, column=1, sticky=tk.EW)
        self.entry_card_height.insert(0, str(settings.get("card_height", 55.0)))

        tk.Label(frame_layout, text="Page Size (WxH mm):").grid(
            row=3, column=0, sticky=tk.W
        )
        self.entry_page_size = tk.Entry(frame_layout)
        self.entry_page_size.grid(row=3, column=1, sticky=tk.EW)
        self.entry_page_size.insert(0, settings.get("page_size", "210x297"))

        tk.Label(frame_layout, text="Margins (Top,Bottom,Left,Right mm):").grid(
            row=4, column=0, sticky=tk.W
        )
        self.entry_margins = tk.Entry(frame_layout)
        self.entry_margins.grid(row=4, column=1, sticky=tk.EW)
        self.entry_margins.insert(0, settings.get("margins", "10,10,10,10"))

        # Section: Font Details
        frame_font = tk.LabelFrame(frame_main, text="Font Details", padx=5, pady=5)
        frame_font.grid(row=2, column=0, sticky=tk.EW, pady=5)
        frame_font.columnconfigure(0, weight=1)
        frame_font.columnconfigure(1, weight=1)

        # Front Font Subframe
        frame_font_front = tk.LabelFrame(frame_font, text="Front", padx=5, pady=5)
        frame_font_front.grid(row=0, column=0, sticky=tk.EW, padx=5)
        frame_font_front.columnconfigure(1, weight=1)

        tk.Label(frame_font_front, text="Font Size:").grid(row=0, column=0, sticky=tk.W)
        self.entry_front_font_size = tk.Entry(frame_font_front)
        self.entry_front_font_size.grid(row=0, column=1, sticky=tk.EW)
        self.entry_front_font_size.insert(0, str(settings.get("front_font_size", 12)))

        tk.Label(frame_font_front, text="Font Family:").grid(row=1, column=0, sticky=tk.W)
        self.front_font_family_var = tk.StringVar(
            value=settings.get("front_font_family", "Helvetica")
        )
        font_families = ["Helvetica", "Times-Roman", "Courier", "Arial"]
        tk.OptionMenu(frame_font_front, self.front_font_family_var, *font_families).grid(
            row=1, column=1, sticky=tk.EW
        )

        tk.Label(frame_font_front, text="Font Style:").grid(row=2, column=0, sticky=tk.W)
        self.front_font_style_var = tk.StringVar(
            value=settings.get("front_font_style", "Normal")
        )
        font_styles = ["Normal", "Bold", "Italic", "BoldItalic"]
        tk.OptionMenu(frame_font_front, self.front_font_style_var, *font_styles).grid(
            row=2, column=1, sticky=tk.EW
        )

        # Back Font Subframe
        frame_font_back = tk.LabelFrame(frame_font, text="Back", padx=5, pady=5)
        frame_font_back.grid(row=0, column=1, sticky=tk.EW, padx=5)
        frame_font_back.columnconfigure(1, weight=1)

        tk.Label(frame_font_back, text="Font Size:").grid(row=0, column=0, sticky=tk.W)
        self.entry_back_font_size = tk.Entry(frame_font_back)
        self.entry_back_font_size.grid(row=0, column=1, sticky=tk.EW)
        self.entry_back_font_size.insert(0, str(settings.get("back_font_size", 12)))

        tk.Label(frame_font_back, text="Font Family:").grid(row=1, column=0, sticky=tk.W)
        self.back_font_family_var = tk.StringVar(
            value=settings.get("back_font_family", "Helvetica")
        )
        tk.OptionMenu(frame_font_back, self.back_font_family_var, *font_families).grid(
            row=1, column=1, sticky=tk.EW
        )

        tk.Label(frame_font_back, text="Font Style:").grid(row=2, column=0, sticky=tk.W)
        self.back_font_style_var = tk.StringVar(
            value=settings.get("back_font_style", "Normal")
        )
        tk.OptionMenu(frame_font_back, self.back_font_style_var, *font_styles).grid(
            row=2, column=1, sticky=tk.EW
        )

        # Section: Colors
        frame_colors = tk.LabelFrame(frame_main, text="Colors", padx=5, pady=5)
        frame_colors.grid(row=3, column=0, sticky=tk.EW, pady=5)
        frame_colors.columnconfigure(0, weight=1)
        frame_colors.columnconfigure(1, weight=1)

        # Front Colors Subframe
        frame_colors_front = tk.LabelFrame(frame_colors, text="Front", padx=5, pady=5)
        frame_colors_front.grid(row=0, column=0, sticky=tk.EW, padx=5)
        frame_colors_front.columnconfigure(1, weight=1)
        frame_colors_front.columnconfigure(2, weight=0)
        frame_colors_front.columnconfigure(3, weight=0)

        tk.Label(frame_colors_front, text="Text Color:").grid(row=0, column=0, sticky=tk.W)
        self.front_text_color_var = tk.StringVar(
            value=settings.get("front_text_color", "#000000")
        )
        self.lbl_front_text_color = tk.Label(
            frame_colors_front,
            textvariable=self.front_text_color_var,
            bg=self.front_text_color_var.get(),
            width=10,
        )
        self.lbl_front_text_color.grid(row=0, column=1, sticky=tk.W)
        btn_front_text_color = tk.Button(
            frame_colors_front,
            text="Pick Text Color",
            command=self.pick_front_text_color,
        )
        btn_front_text_color.grid(row=0, column=2, sticky=tk.W)

        tk.Label(frame_colors_front, text="Background Color:").grid(
            row=1, column=0, sticky=tk.W
        )
        self.front_background_color_var = tk.StringVar(
            value=settings.get("front_background_color", "#FFFFFF")
        )
        self.lbl_front_background_color = tk.Label(
            frame_colors_front,
            textvariable=self.front_background_color_var,
            bg=self.front_background_color_var.get(),
            width=10,
        )
        self.lbl_front_background_color.grid(row=1, column=1, sticky=tk.W)
        btn_front_background_color = tk.Button(
            frame_colors_front,
            text="Pick Background Color",
            command=self.pick_front_background_color,
        )
        btn_front_background_color.grid(row=1, column=2, sticky=tk.W)

        self.front_color_bar_top_var = tk.BooleanVar(
            value=settings.get("front_color_bar_top", False)
        )
        tk.Checkbutton(
            frame_colors_front,
            text="Add Color Bar Top",
            variable=self.front_color_bar_top_var,
        ).grid(row=2, column=0, sticky=tk.W)
        self.front_color_bar_top_color = tk.StringVar(
            value=settings.get("front_color_bar_top_color", "#FF0000")
        )
        self.lbl_front_top_color = tk.Label(
            frame_colors_front,
            textvariable=self.front_color_bar_top_color,
            bg=self.front_color_bar_top_color.get(),
            width=10,
        )
        self.lbl_front_top_color.grid(row=2, column=1, sticky=tk.W)
        btn_front_top_color = tk.Button(
            frame_colors_front,
            text="Pick Top Color",
            command=self.pick_front_top_color,
        )
        btn_front_top_color.grid(row=2, column=2, sticky=tk.W)

        self.front_color_bar_bottom_var = tk.BooleanVar(
            value=settings.get("front_color_bar_bottom", False)
        )
        tk.Checkbutton(
            frame_colors_front,
            text="Add Color Bar Bottom",
            variable=self.front_color_bar_bottom_var,
        ).grid(row=3, column=0, sticky=tk.W)
        self.front_color_bar_bottom_color = tk.StringVar(
            value=settings.get("front_color_bar_bottom_color", "#0000FF")
        )
        self.lbl_front_bottom_color = tk.Label(
            frame_colors_front,
            textvariable=self.front_color_bar_bottom_color,
            bg=self.front_color_bar_bottom_color.get(),
            width=10,
        )
        self.lbl_front_bottom_color.grid(row=3, column=1, sticky=tk.W)
        btn_front_bottom_color = tk.Button(
            frame_colors_front,
            text="Pick Bottom Color",
            command=self.pick_front_bottom_color,
        )
        btn_front_bottom_color.grid(row=3, column=2, sticky=tk.W)

        # Back Colors Subframe
        frame_colors_back = tk.LabelFrame(frame_colors, text="Back", padx=5, pady=5)
        frame_colors_back.grid(row=0, column=1, sticky=tk.EW, padx=5)
        frame_colors_back.columnconfigure(1, weight=1)
        frame_colors_back.columnconfigure(2, weight=0)
        frame_colors_back.columnconfigure(3, weight=0)

        tk.Label(frame_colors_back, text="Text Color:").grid(row=0, column=0, sticky=tk.W)
        self.back_text_color_var = tk.StringVar(
            value=settings.get("back_text_color", "#000000")
        )
        self.lbl_back_text_color = tk.Label(
            frame_colors_back,
            textvariable=self.back_text_color_var,
            bg=self.back_text_color_var.get(),
            width=10,
        )
        self.lbl_back_text_color.grid(row=0, column=1, sticky=tk.W)
        btn_back_text_color = tk.Button(
            frame_colors_back,
            text="Pick Text Color",
            command=self.pick_back_text_color,
        )
        btn_back_text_color.grid(row=0, column=2, sticky=tk.W)

        tk.Label(frame_colors_back, text="Background Color:").grid(
            row=1, column=0, sticky=tk.W
        )
        self.back_background_color_var = tk.StringVar(
            value=settings.get("back_background_color", "#FFFFFF")
        )
        self.lbl_back_background_color = tk.Label(
            frame_colors_back,
            textvariable=self.back_background_color_var,
            bg=self.back_background_color_var.get(),
            width=10,
        )
        self.lbl_back_background_color.grid(row=1, column=1, sticky=tk.W)
        btn_back_background_color = tk.Button(
            frame_colors_back,
            text="Pick Background Color",
            command=self.pick_back_background_color,
        )
        btn_back_background_color.grid(row=1, column=2, sticky=tk.W)

        self.back_color_bar_top_var = tk.BooleanVar(
            value=settings.get("back_color_bar_top", False)
        )
        tk.Checkbutton(
            frame_colors_back,
            text="Add Color Bar Top",
            variable=self.back_color_bar_top_var,
        ).grid(row=2, column=0, sticky=tk.W)
        self.back_color_bar_top_color = tk.StringVar(
            value=settings.get("back_color_bar_top_color", "#FF0000")
        )
        self.lbl_back_top_color = tk.Label(
            frame_colors_back,
            textvariable=self.back_color_bar_top_color,
            bg=self.back_color_bar_top_color.get(),
            width=10,
        )
        self.lbl_back_top_color.grid(row=2, column=1, sticky=tk.W)
        btn_back_top_color = tk.Button(
            frame_colors_back,
            text="Pick Top Color",
            command=self.pick_back_top_color,
        )
        btn_back_top_color.grid(row=2, column=2, sticky=tk.W)

        self.back_color_bar_bottom_var = tk.BooleanVar(
            value=settings.get("back_color_bar_bottom", False)
        )
        tk.Checkbutton(
            frame_colors_back,
            text="Add Color Bar Bottom",
            variable=self.back_color_bar_bottom_var,
        ).grid(row=3, column=0, sticky=tk.W)
        self.back_color_bar_bottom_color = tk.StringVar(
            value=settings.get("back_color_bar_bottom_color", "#0000FF")
        )
        self.lbl_back_bottom_color = tk.Label(
            frame_colors_back,
            textvariable=self.back_color_bar_bottom_color,
            bg=self.back_color_bar_bottom_color.get(),
            width=10,
        )
        self.lbl_back_bottom_color.grid(row=3, column=1, sticky=tk.W)
        btn_back_bottom_color = tk.Button(
            frame_colors_back,
            text="Pick Bottom Color",
            command=self.pick_back_bottom_color,
        )
        btn_back_bottom_color.grid(row=3, column=2, sticky=tk.W)

        # Section: Manipulators
        frame_manipulators = tk.LabelFrame(
            frame_main, text="Manipulators", padx=5, pady=5
        )
        frame_manipulators.grid(row=4, column=0, sticky=tk.EW, pady=5)
        frame_manipulators.columnconfigure(1, weight=1)

        self.truncate_var = tk.BooleanVar(value=settings.get("truncate", False))
        tk.Checkbutton(
            frame_manipulators,
            text="Truncate overflow text",
            variable=self.truncate_var,
        ).grid(row=0, column=0, sticky=tk.W)

        self.use_qty_column_var = tk.BooleanVar(
            value=settings.get("use_qty_column", True)
        )
        tk.Checkbutton(
            frame_manipulators,
            text="Various amounts from 'qty' col",
            variable=self.use_qty_column_var,
        ).grid(row=1, column=0, sticky=tk.W)

        tk.Label(frame_manipulators, text="Multiple (all cards):").grid(
            row=2, column=0, sticky=tk.W
        )
        self.entry_quantity_multiplier = tk.Entry(frame_manipulators, width=5)
        self.entry_quantity_multiplier.grid(row=2, column=1, sticky=tk.W)
        self.entry_quantity_multiplier.insert(
            0, str(settings.get("quantity_multiplier", 1))
        )

        tk.Label(frame_manipulators, text="Flip Mode:").grid(
            row=3, column=0, sticky=tk.W
        )
        self.flip_mode_var = tk.StringVar(value=settings.get("flip_mode", "long"))
        tk.Radiobutton(
            frame_manipulators,
            text="Flip on Long Edge",
            variable=self.flip_mode_var,
            value="long",
        ).grid(row=3, column=1, sticky=tk.W)
        tk.Radiobutton(
            frame_manipulators,
            text="Flip on Short Edge",
            variable=self.flip_mode_var,
            value="short",
        ).grid(row=3, column=2, sticky=tk.W)

        # Buttons frame
        frame_buttons = tk.Frame(self.scrollable_frame)
        frame_buttons.pack(fill=tk.X, padx=10, pady=10)
        frame_buttons.columnconfigure(0, weight=1)
        frame_buttons.columnconfigure(1, weight=0)
        frame_buttons.columnconfigure(2, weight=1)
        frame_buttons.columnconfigure(3, weight=0)
        frame_buttons.columnconfigure(4, weight=1)

        btn_start = tk.Button(frame_buttons, text="Start", command=self.start_process)
        btn_start.grid(row=0, column=1, sticky=tk.EW)
        btn_exit = tk.Button(frame_buttons, text="Exit", command=self.quit)
        btn_exit.grid(row=0, column=3, sticky=tk.EW)

        # Initialize column entries state
        self.toggle_column_entries()
        logger.info("GUI initialization complete")

    def _on_frame_configure(self, event):
        """Update the scroll region to encompass the scrollable frame."""
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        """Adjust the canvas window width to match the canvas width."""
        self.canvas.itemconfig(self.canvas_frame, width=event.width)

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling for the canvas."""
        if event.num == 4 or event.delta > 0:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5 or event.delta < 0:
            self.canvas.yview_scroll(1, "units")

    def toggle_column_entries(self):
        """Enable or disable front/back column entry fields based on autofill checkbox and update values."""
        autofill_enabled = self.autofill_columns_var.get()
        state = tk.DISABLED if autofill_enabled else tk.NORMAL
        logger.info(f"Setting column entries state to {state}")

        # Always clear entries to avoid stale values
        self.entry_front.delete(0, tk.END)
        self.entry_back.delete(0, tk.END)

        # Temporarily enable entries to allow updates
        self.entry_front.config(state=tk.NORMAL)
        self.entry_back.config(state=tk.NORMAL)

        if autofill_enabled and self.entry_file.get():
            self.autofill_column_names(self.entry_file.get())
        else:
            # Restore saved or default values if autofill is disabled
            self.entry_front.insert(0, settings.get("front_column", ""))
            self.entry_back.insert(0, settings.get("back_column", ""))
            logger.info(
                f"Restored front column to '{settings.get('front_column', '')}', back column to '{settings.get('back_column', '')}'"
            )

        # Apply the desired state
        self.entry_front.config(state=state)
        self.entry_back.config(state=state)

        # Force GUI update
        self.entry_front.update()
        self.entry_back.update()
        self.update_idletasks()
        logger.info(
            f"Column entries updated, front: '{self.entry_front.get()}', back: '{self.entry_back.get()}'"
        )

    def browse_file(self):
        """Open a file dialog and autofill column names if enabled, clearing previous values."""
        logger.info("Opening file dialog for input file selection")
        file_path = filedialog.askopenfilename(
            filetypes=[("Spreadsheets", "*.csv;*.xlsx;*.xls;*.ods"), ("All files", "*")]
        )
        if file_path:
            logger.info(f"Selected input file: {file_path}")
            self.entry_file.delete(0, tk.END)
            self.entry_file.insert(0, file_path)
            default_output = os.path.splitext(file_path)[0] + ".pdf"
            self.entry_output.delete(0, tk.END)
            self.entry_output.insert(0, default_output)

            # Clear and update column entries
            self.entry_front.delete(0, tk.END)
            self.entry_back.delete(0, tk.END)
            if self.autofill_columns_var.get():
                self.entry_front.config(state=tk.NORMAL)
                self.entry_back.config(state=tk.NORMAL)
                self.autofill_column_names(file_path)
                self.entry_front.config(state=tk.DISABLED)
                self.entry_back.config(state=tk.DISABLED)
            else:
                self.entry_front.insert(0, settings.get("front_column", ""))
                self.entry_back.insert(0, settings.get("back_column", ""))
                self.entry_front.config(state=tk.NORMAL)
                self.entry_back.config(state=tk.NORMAL)

            # Force GUI update
            self.entry_front.update()
            self.entry_back.update()
            self.update_idletasks()
            logger.info(
                f"After file selection, front: '{self.entry_front.get()}', back: '{self.entry_back.get()}'"
            )
        else:
            logger.info("No input file selected")

    def autofill_column_names(self, file_path):
        """Read the first two column headers from the file and set them as front/back column names."""
        logger.info(f"Attempting to autofill column names from {file_path}")
        if not os.path.exists(file_path):
            logger.error(f"File does not exist: {file_path}")
            messagebox.showerror("Error", f"File does not exist: {file_path}")
            return

        try:
            headers = []
            if file_path.lower().endswith(".csv"):
                df = pd.read_csv(file_path, nrows=0, encoding="utf-8")
                headers = df.columns.tolist()
                logger.debug(f"CSV headers: {headers}")
            elif file_path.lower().endswith((".xlsx", ".xls")):
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                headers = [
                    cell.value for cell in next(ws.rows) if cell.value is not None
                ]
                headers = [str(h) if h is not None else "Unnamed" for h in headers]
                logger.debug(f"Excel headers: {headers}")
                wb.close()
            elif file_path.lower().endswith(".ods"):
                doc = odf.opendocument.load(file_path)
                sheet = doc.spreadsheet.getElementsByType(Table)[0]
                header_row = sheet.getElementsByType(odf.table.TableRow)[0]
                headers = []
                for cell in header_row.getElementsByType(odf.table.TableCell):
                    paragraphs = cell.getElementsByType(P)
                    text = (
                        "".join(p.firstChild.data for p in paragraphs if p.firstChild)
                        if paragraphs
                        else ""
                    )
                    headers.append(text.strip() or "Unnamed")
                logger.debug(f"ODS headers: {headers}")
            else:
                logger.error("Unsupported file format")
                messagebox.showerror(
                    "Error", "Unsupported file format. Please use CSV, Excel, or ODS."
                )
                return

            # Sanitize headers
            headers = [str(h).strip() if h else "Unnamed" for h in headers]
            logger.info(f"Sanitized headers: {headers}")

            # Clear existing entries
            self.entry_front.delete(0, tk.END)
            self.entry_back.delete(0, tk.END)

            # Assign headers to entry fields
            if headers:
                self.entry_front.insert(0, headers[0])
                logger.info(f"Set front column to '{headers[0]}'")
            else:
                self.entry_front.insert(0, "")
                logger.warning("No headers found in file")
                messagebox.showwarning(
                    "Warning", "No columns found in file to autofill."
                )

            if len(headers) >= 2:
                self.entry_back.insert(0, headers[1])
                logger.info(f"Set back column to '{headers[1]}'")
            else:
                self.entry_back.insert(0, "")
                logger.warning("Not enough columns for back column autofill")
                if headers:
                    messagebox.showwarning(
                        "Warning", "Only one column found; back column set to empty."
                    )

            # Force GUI update
            self.entry_front.update()
            self.entry_back.update()
            self.update_idletasks()
            logger.info(
                f"Autofill completed, front: '{self.entry_front.get()}', back: '{self.entry_back.get()}'"
            )

        except Exception as e:
            logger.error(f"Failed to autofill column names: {e}")
            messagebox.showerror("Error", f"Failed to autofill column names: {e}")

    def browse_output_file(self):
        """Open a file dialog to select the output PDF file."""
        logger.info("Opening file dialog for output file selection")
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")]
        )
        if file_path:
            logger.info(f"Selected output file: {file_path}")
            self.entry_output.delete(0, tk.END)
            self.entry_output.insert(0, file_path)

    def pick_front_text_color(self):
        """Open a color picker dialog for the front text color."""
        logger.info("Opening color picker for front text color")
        color = colorchooser.askcolor(title="Pick Front Text Color")[1]
        if color:
            self.front_text_color_var.set(color)
            self.lbl_front_text_color.config(bg=color)
            logger.info(f"Selected front text color: {color}")

    def pick_front_background_color(self):
        """Open a color picker dialog for the front background color."""
        logger.info("Opening color picker for front background color")
        color = colorchooser.askcolor(title="Pick Front Background Color")[1]
        if color:
            self.front_background_color_var.set(color)
            self.lbl_front_background_color.config(bg=color)
            logger.info(f"Selected front background color: {color}")

    def pick_front_top_color(self):
        """Open a color picker dialog for the front top color bar."""
        logger.info("Opening color picker for front top bar")
        color = colorchooser.askcolor(title="Pick Front Top Bar Color")[1]
        if color:
            self.front_color_bar_top_color.set(color)
            self.lbl_front_top_color.config(bg=color)
            logger.info(f"Selected front top bar color: {color}")

    def pick_front_bottom_color(self):
        """Open a color picker dialog for the front bottom color bar."""
        logger.info("Opening color picker for front bottom bar")
        color = colorchooser.askcolor(title="Pick Front Bottom Bar Color")[1]
        if color:
            self.front_color_bar_bottom_color.set(color)
            self.lbl_front_bottom_color.config(bg=color)
            logger.info(f"Selected front bottom bar color: {color}")

    def pick_back_text_color(self):
        """Open a color picker dialog for the back text color."""
        logger.info("Opening color picker for back text color")
        color = colorchooser.askcolor(title="Pick Back Text Color")[1]
        if color:
            self.back_text_color_var.set(color)
            self.lbl_back_text_color.config(bg=color)
            logger.info(f"Selected back text color: {color}")

    def pick_back_background_color(self):
        """Open a color picker dialog for the back background color."""
        logger.info("Opening color picker for back background color")
        color = colorchooser.askcolor(title="Pick Back Background Color")[1]
        if color:
            self.back_background_color_var.set(color)
            self.lbl_back_background_color.config(bg=color)
            logger.info(f"Selected back background color: {color}")

    def pick_back_top_color(self):
        """Open a color picker dialog for the back top color bar."""
        logger.info("Opening color picker for back top bar")
        color = colorchooser.askcolor(title="Pick Back Top Bar Color")[1]
        if color:
            self.back_color_bar_top_color.set(color)
            self.lbl_back_top_color.config(bg=color)
            logger.info(f"Selected back top bar color: {color}")

    def pick_back_bottom_color(self):
        """Open a color picker dialog for the back bottom color bar."""
        logger.info("Opening color picker for back bottom bar")
        color = colorchooser.askcolor(title="Pick Back Bottom Bar Color")[1]
        if color:
            self.back_color_bar_bottom_color.set(color)
            self.lbl_back_bottom_color.config(bg=color)
            logger.info(f"Selected back bottom bar color: {color}")

    def start_process(self):
        """
        Validate inputs, save settings, load data, validate 'qty' column if used, and generate the PDF.
        """
        logger.info("Starting PDF generation process")
        settings["file_path"] = self.entry_file.get()
        settings["output_file"] = self.entry_output.get()
        settings["front_column"] = self.entry_front.get()
        settings["back_column"] = self.entry_back.get()
        settings["use_qty_column"] = self.use_qty_column_var.get()
        settings["autofill_columns"] = self.autofill_columns_var.get()
        try:
            settings["cards_per_row"] = int(self.entry_cards_per_row.get())
            settings["card_width"] = float(self.entry_card_width.get())
            settings["card_height"] = float(self.entry_card_height.get())
            settings["page_size"] = self.entry_page_size.get()
            settings["margins"] = self.entry_margins.get()
            settings["front_font_size"] = float(self.entry_front_font_size.get())
            settings["back_font_size"] = float(self.entry_back_font_size.get())
            multiplier = self.entry_quantity_multiplier.get().strip()
            if multiplier and len(multiplier) > 4:
                raise ValueError("Quantity multiplier cannot exceed 4 digits")
            settings["quantity_multiplier"] = int(multiplier or 1)
            if settings["quantity_multiplier"] <= 0:
                raise ValueError("Quantity multiplier must be a positive integer")
        except ValueError as e:
            logger.error(f"Invalid input in numeric fields: {e}")
            messagebox.showerror("Error", f"Invalid input in numeric fields: {e}")
            return
        settings["front_color_bar_top"] = self.front_color_bar_top_var.get()
        settings["front_color_bar_bottom"] = self.front_color_bar_bottom_var.get()
        settings["front_color_bar_top_color"] = self.front_color_bar_top_color.get()
        settings["front_color_bar_bottom_color"] = self.front_color_bar_bottom_color.get()
        settings["front_background_color"] = self.front_background_color_var.get()
        settings["back_color_bar_top"] = self.back_color_bar_top_var.get()
        settings["back_color_bar_bottom"] = self.back_color_bar_bottom_var.get()
        settings["back_color_bar_top_color"] = self.back_color_bar_top_color.get()
        settings["back_color_bar_bottom_color"] = self.back_color_bar_bottom_color.get()
        settings["back_background_color"] = self.back_background_color_var.get()
        settings["truncate"] = self.truncate_var.get()
        settings["flip_mode"] = self.flip_mode_var.get()
        settings["front_font_family"] = self.front_font_family_var.get()
        settings["front_font_style"] = self.front_font_style_var.get()
        settings["front_text_color"] = self.front_text_color_var.get()
        settings["back_font_family"] = self.back_font_family_var.get()
        settings["back_font_style"] = self.back_font_style_var.get()
        settings["back_text_color"] = self.back_text_color_var.get()

        try:
            with open(SETTINGS_FILE, "w") as f:
                json.dump(settings, f, indent=4)
                logger.info(f"Settings saved to {SETTINGS_FILE}")
        except Exception as e:
            logger.error(f"Failed to save settings: {e}")
            messagebox.showerror("Error", f"Failed to save settings: {e}")
            return

        file_path = settings["file_path"]
        if not file_path:
            logger.error("No input file selected")
            messagebox.showerror("Error", "Please select an input file.")
            return

        logger.info(f"Reading input file: {file_path}")
        try:
            if file_path.lower().endswith(".csv"):
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
                content = content.replace("\\n", "\n")
                data = pd.read_csv(
                    io.StringIO(content),
                    quoting=csv.QUOTE_ALL,
                    keep_default_na=False,
                    lineterminator="\n",
                )
            elif file_path.lower().endswith((".xlsx", ".xls")):
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                data_rows = []
                headers = [cell.value for cell in ws[1] if cell.value is not None]
                for row in ws.iter_rows(min_row=2):
                    row_data = [
                        cell.value if cell.value is not None else "" for cell in row
                    ]
                    row_data = [
                        ("\n".join(str(val).splitlines()) if val else "")
                        for val in row_data
                    ]
                    data_rows.append(row_data)
                data = pd.DataFrame(data_rows, columns=headers)
            elif file_path.lower().endswith(".ods"):
                doc = odf.opendocument.load(file_path)
                data_rows = []
                sheet = doc.spreadsheet.getElementsByType(Table)[0]
                header_row = sheet.getElementsByType(odf.table.TableRow)[0]
                headers = []
                for cell in header_row.getElementsByType(odf.table.TableCell):
                    paragraphs = cell.getElementsByType(P)
                    text = "\n".join(
                        p.firstChild.data for p in paragraphs if p.firstChild
                    )
                    headers.append(text.strip() or "Unnamed")
                expected_cols = len(headers)
                for row in sheet.getElementsByType(odf.table.TableRow)[1:]:
                    row_data = []
                    for cell in row.getElementsByType(odf.table.TableCell):
                        paragraphs = cell.getElementsByType(P)
                        text = "\n".join(
                            p.firstChild.data for p in paragraphs if p.firstChild
                        )
                        row_data.append(text.strip() or "")
                    while len(row_data) < expected_cols:
                        row_data.append("")
                    if len(row_data) > expected_cols:
                        row_data = row_data[:expected_cols]
                    data_rows.append(row_data)
                data = pd.DataFrame(data_rows, columns=headers)
                logger.warning("ODS file processed; verify newlines in output")
            else:
                logger.error("Unsupported file format")
                messagebox.showerror(
                    "Error", "Unsupported file format. Please use CSV, Excel, or ODS."
                )
                return
            logger.info(f"Read {len(data)} rows with {len(data.columns)} columns")
        except Exception as e:
            logger.error(f"Failed to read file: {e}")
            messagebox.showerror("Error", f"Failed to read file: {e}")
            return

        missing = []
        for col in [settings["front_column"], settings["back_column"]]:
            if col and col not in data.columns:
                missing.append(col)
        if settings["use_qty_column"] and "qty" not in data.columns:
            missing.append("qty")
        if missing:
            logger.error(f"Missing columns: {', '.join(missing)}")
            messagebox.showerror(
                "Error", f"Missing columns in data: {', '.join(missing)}"
            )
            return

        if settings["use_qty_column"]:
            try:
                data["qty"] = data["qty"].fillna(1)
                invalid_rows = []
                qty_values = pd.to_numeric(data["qty"], errors="coerce")
                non_numeric_rows = data.index[qty_values.isna()].tolist()
                for row in non_numeric_rows:
                    invalid_rows.append(
                        f"Row {row + 2}: Non-numeric value '{data['qty'].iloc[row]}'"
                    )
                qty_values = qty_values.fillna(1)
                non_integer_rows = qty_values.index[
                    qty_values != qty_values.astype(int)
                ].tolist()
                for row in non_integer_rows:
                    invalid_rows.append(
                        f"Row {row + 2}: Non-integer value '{data['qty'].iloc[row]}'"
                    )
                non_positive_rows = qty_values.index[qty_values <= 0].tolist()
                for row in non_positive_rows:
                    invalid_rows.append(
                        f"Row {row + 2}: Non-positive value '{data['qty'].iloc[row]}'"
                    )
                data["qty"] = qty_values.where(
                    (qty_values > 0) & (qty_values == qty_values.astype(int)), 1
                ).astype(int)
                if invalid_rows:
                    messagebox.showwarning(
                        "Warning",
                        "Issues found in 'qty' column (defaulted to 1):\n"
                        + "\n".join(invalid_rows),
                    )
            except Exception as e:
                logger.error(f"Error validating 'qty' column: {e}")
                messagebox.showerror("Error", f"Error validating 'qty' column: {e}")
                return

        try:
            self.generate_flashcard_pdf(data, settings)
            messagebox.showinfo(
                "Success", f"PDF generated as {settings['output_file']}"
            )
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            messagebox.showerror("Error", f"PDF generation failed: {e}")

    def generate_flashcard_pdf(self, data, settings):
        """
        Generate a flashcard PDF with alternating front/back pages, centered on the page, with variable quantities.

        Args:
            data (pandas.DataFrame): Input data with front/back columns and optional 'qty' column.
            settings (dict): Configuration settings for PDF generation.
        """
        logger.info("Starting PDF generation")
        try:
            page_w, page_h = map(float, settings["page_size"].split("x"))
            margins = list(map(float, settings["margins"].split(",")))
            margin_top, margin_bottom, margin_left, margin_right = margins
        except ValueError as e:
            logger.error(f"Invalid page size or margins format: {e}")
            raise ValueError(f"Invalid page size or margins format: {e}")

        cards_per_row = int(settings["cards_per_row"])
        card_width = float(settings["card_width"])
        card_height = float(settings["card_height"])
        front_font_size = float(settings["front_font_size"])
        back_font_size = float(settings["back_font_size"])
        front_font_family = settings["front_font_family"]
        front_font_style = settings["front_font_style"]
        back_font_family = settings["back_font_family"]
        back_font_style = settings["back_font_style"]
        front_text_color = settings["front_text_color"]
        back_text_color = settings["back_text_color"]
        front_background_color = settings["front_background_color"]
        back_background_color = settings["back_background_color"]
        use_qty_column = settings["use_qty_column"]
        quantity_multiplier = int(settings["quantity_multiplier"])

        font_map = {
            ("Helvetica", "Normal"): "Helvetica",
            ("Helvetica", "Bold"): "Helvetica-Bold",
            ("Helvetica", "Italic"): "Helvetica-Oblique",
            ("Helvetica", "BoldItalic"): "Helvetica-BoldOblique",
            ("Times-Roman", "Normal"): "Times-Roman",
            ("Times-Roman", "Bold"): "Times-Bold",
            ("Times-Roman", "Italic"): "Times-Italic",
            ("Times-Roman", "BoldItalic"): "Times-BoldItalic",
            ("Courier", "Normal"): "Courier",
            ("Courier", "Bold"): "Courier-Bold",
            ("Courier", "Italic"): "Courier-Oblique",
            ("Courier", "BoldItalic"): "Courier-BoldOblique",
            ("Arial", "Normal"): "Helvetica",
            ("Arial", "Bold"): "Helvetica-Bold",
            ("Arial", "Italic"): "Helvetica-Oblique",
            ("Arial", "BoldItalic"): "Helvetica-BoldOblique",
        }
        front_font_name = font_map.get((front_font_family, front_font_style), "Helvetica")
        back_font_name = font_map.get((back_font_family, back_font_style), "Helvetica")

        page_size_pt = (mm(page_w), mm(page_h))
        card_width_pt = mm(card_width)
        card_height_pt = mm(card_height)

        cards_per_col = int((page_h - margin_top - margin_bottom) // card_height)
        cards_per_page = cards_per_row * cards_per_col

        grid_width = cards_per_row * card_width
        grid_height = cards_per_col * card_height
        if (
            grid_width + margin_left + margin_right > page_w
            or grid_height + margin_top + margin_bottom > page_h
        ):
            raise ValueError("Card grid with margins is too large for the page size.")
        if cards_per_col <= 0 or cards_per_row <= 0:
            raise ValueError("Card size or margins too large for page size.")

        leftover_width = page_w - grid_width - margin_left - margin_right
        leftover_height = page_h - grid_height - margin_top - margin_bottom
        offset_left = margin_left + leftover_width / 2
        offset_top = margin_top + leftover_height / 2
        offset_left_pt = mm(offset_left)
        offset_top_pt = mm(offset_top)

        output_file = settings.get("output_file", "flashcards.pdf")
        c = canvas.Canvas(output_file, pagesize=page_size_pt)
        front_line_height = front_font_size * 1.2
        back_line_height = back_font_size * 1.2

        card_indices = []
        for idx in range(len(data)):
            qty = int(data["qty"].iloc[idx]) if use_qty_column else 1
            qty = qty * quantity_multiplier
            card_indices.extend([idx] * qty)

        num_cards = len(card_indices)
        num_pages = math.ceil(num_cards / cards_per_page)
        flip_mode = settings.get("flip_mode", "long")

        def draw_color_bar(x, y, width, height, color_hex):
            c.setFillColor(HexColor(color_hex))
            c.rect(x, y, width, height, fill=1, stroke=0)

        for page in range(num_pages):
            page_start = page * cards_per_page
            page_end = min(page_start + cards_per_page, num_cards)
            page_indices = card_indices[page_start:page_end]
            while len(page_indices) < cards_per_page:
                page_indices.append(None)

            # Front page
            for pos_on_page, data_idx in enumerate(page_indices):
                card_row = pos_on_page // cards_per_row
                card_col = pos_on_page % cards_per_row
                x = offset_left_pt + card_col * card_width_pt
                y = page_size_pt[1] - offset_top_pt - (card_row + 1) * card_height_pt

                if data_idx is not None:
                    row = data.iloc[data_idx]
                    # Draw background
                    c.setFillColor(HexColor(front_background_color))
                    c.rect(x, y, card_width_pt, card_height_pt, fill=1, stroke=0)
                    # Draw color bars
                    if settings["front_color_bar_top"]:
                        draw_color_bar(
                            x,
                            y + card_height_pt - mm(5),
                            card_width_pt,
                            mm(5),
                            settings["front_color_bar_top_color"],
                        )
                    if settings["front_color_bar_bottom"]:
                        draw_color_bar(
                            x,
                            y,
                            card_width_pt,
                            mm(5),
                            settings["front_color_bar_bottom_color"],
                        )

                    text = str(row.get(settings["front_column"], ""))
                    max_text_height = card_height_pt - mm(8)
                    if settings["front_color_bar_top"]:
                        max_text_height -= mm(5)
                    if settings["front_color_bar_bottom"]:
                        max_text_height -= mm(5)
                    max_lines = int(max_text_height // front_line_height)
                    lines, overflowed = wrap_text(
                        text,
                        front_font_name,
                        front_font_size,
                        card_width_pt - mm(8),
                        max_lines,
                        settings.get("truncate", False),
                    )
                    if overflowed and not settings.get("truncate", False):
                        raise Exception(
                            f"Text does not fit on card at row {data_idx+1}. Enable 'Truncate' or edit your data."
                        )

                    c.setFont(front_font_name, front_font_size)
                    c.setFillColor(HexColor(front_text_color))
                    text_height = len(lines) * front_line_height
                    text_y = y + (card_height_pt - text_height) / 2
                    for lidx, line in enumerate(lines):
                        c.drawCentredString(
                            x + card_width_pt / 2,
                            text_y + text_height - (lidx + 1) * front_line_height,
                            line,
                        )

            draw_cut_lines(
                c,
                page_size_pt[0],
                page_size_pt[1],
                offset_left_pt,
                offset_top_pt,
                card_width_pt,
                card_height_pt,
                cards_per_row,
                cards_per_col,
            )
            c.showPage()

            # Back page
            back_order = reorder_for_back(
                page_indices, cards_per_row, cards_per_col, flip_mode
            )
            for pos_on_page, data_idx in enumerate(back_order):
                card_row = pos_on_page // cards_per_row
                card_col = pos_on_page % cards_per_row
                x = offset_left_pt + card_col * card_width_pt
                y = page_size_pt[1] - offset_top_pt - (card_row + 1) * card_height_pt

                if data_idx is not None:
                    row = data.iloc[data_idx]
                    # Draw background
                    c.setFillColor(HexColor(back_background_color))
                    c.rect(x, y, card_width_pt, card_height_pt, fill=1, stroke=0)
                    # Draw color bars
                    if settings["back_color_bar_top"]:
                        draw_color_bar(
                            x,
                            y + card_height_pt - mm(5),
                            card_width_pt,
                            mm(5),
                            settings["back_color_bar_top_color"],
                        )
                    if settings["back_color_bar_bottom"]:
                        draw_color_bar(
                            x,
                            y,
                            card_width_pt,
                            mm(5),
                            settings["back_color_bar_bottom_color"],
                        )

                    text = str(row.get(settings["back_column"], ""))
                    max_text_height = card_height_pt - mm(8)
                    if settings["back_color_bar_top"]:
                        max_text_height -= mm(5)
                    if settings["back_color_bar_bottom"]:
                        max_text_height -= mm(5)
                    max_lines = int(max_text_height // back_line_height)
                    lines, overflowed = wrap_text(
                        text,
                        back_font_name,
                        back_font_size,
                        card_width_pt - mm(8),
                        max_lines,
                        settings.get("truncate", False),
                    )
                    if overflowed and not settings.get("truncate", False):
                        raise Exception(
                            f"Back text does not fit on card at row {data_idx+1}. Enable 'Truncate' or edit your data."
                        )

                    c.setFont(back_font_name, back_font_size)
                    c.setFillColor(HexColor(back_text_color))
                    text_height = len(lines) * back_line_height
                    text_y = y + (card_height_pt - text_height) / 2
                    for lidx, line in enumerate(lines):
                        c.drawCentredString(
                            x + card_width_pt / 2,
                            text_y + text_height - (lidx + 1) * back_line_height,
                            line,
                        )

            draw_cut_lines(
                c,
                page_size_pt[0],
                page_size_pt[1],
                offset_left_pt,
                offset_top_pt,
                card_width_pt,
                card_height_pt,
                cards_per_row,
                cards_per_col,
            )
            c.showPage()

        c.save()


if __name__ == "__main__":
    """
    Entry point for the Flashcard PDF Generator application.
    """
    logger.info("Starting Flashcard PDF Generator application")
    app = FlashcardApp()
    app.mainloop()
    logger.info("Application closed")